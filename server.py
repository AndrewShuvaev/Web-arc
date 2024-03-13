from flask import Flask, jsonify, request, send_file
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
UPLOAD_FOLDER = 'uploaded_files'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Файл не выбран'})

    filename = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filename)
    return jsonify({'message': 'Файл успешно загружен'})

@app.route('/get_files', methods=['GET'])
def get_files():
    files = os.listdir(app.config['UPLOAD_FOLDER'])
    return jsonify({'files': files})

@app.route('/excel/<filename>', methods=['GET', 'POST', 'DELETE'])
def excel_operations(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    if request.method == 'GET':
        workbook = load_workbook(file_path)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)
        return jsonify({'data': data})

    elif request.method == 'POST':
        data = request.json.get('data')
        if data:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            sheet.append(data)
            workbook.save(file_path)
            return jsonify({'message': 'Данные добавлены'})
        else:
            return jsonify({'error': 'Данные не предоставлены'})

    elif request.method == 'DELETE':
        row_index = request.json.get('row_index')
        if row_index:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            sheet.delete_rows(row_index + 1)  
            workbook.save(file_path)
            return jsonify({'message': 'Строка успешно удалена'})
        else:
            return jsonify({'error': 'Ошибка'})

@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    return send_file(file_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
